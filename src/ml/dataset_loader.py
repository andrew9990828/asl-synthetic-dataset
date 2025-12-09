# ======================================================================
#  dataset_loader.py
#  Synthetic Dataset Generator – PyTorch Dataset Loader
#
#  Author: Andrew Bieber
#
#  Description:
#      Defines a custom PyTorch Dataset class for loading synthetic
#      ASL-like abstract images and their associated labels + distance
#      values stored in labels.xlsx.
#
#      Each sample consists of:
#        - RGB image (resized to 128×128, normalized to [0,1])
#        - A class index (0–25 for A–Z)
#        - A distance regression target (float)
#
#  This loader is designed to mimic real ML research pipelines where
#  metadata is stored separately from image files.
# ======================================================================

import os
import torch
from torch.utils.data import Dataset
from PIL import Image
from openpyxl import load_workbook


# ======================================================================
#  ASLDataset
# ======================================================================
"""
Custom PyTorch Dataset for loading the synthetic ASL abstract dataset.

Constructor Args:
    root (str): Path to the dataset directory. Expects:
                - images/ folder containing generated images
                - labels.xlsx with columns:
                    [image_path, letter_label, distance_value]

Behavior:
    - Reads all dataset metadata from the excel sheet.
    - Builds an internal list of samples.
    - Maps letters A–Z → indices 0–25.
    - Loads, resizes, and converts images into normalized PyTorch tensors.
"""
class ASLDataset(Dataset):

    # ------------------------------------------------------------------
    # __init__
    # ------------------------------------------------------------------
    """
    Loads labels.xlsx, parses metadata, and builds the sample list.

    Steps:
        1. Build absolute path to labels.xlsx.
        2. Load the Excel workbook with openpyxl.
        3. Iterate through rows (skipping header) and extract:
           - relative image path
           - letter label
           - distance value
        4. Store the parsed entries inside self.samples.
        5. Create a dictionary mapping letters → integer indices.
    """
    def __init__(self, root):
        self.root = root
        xl_path = os.path.join(root, "labels.xlsx")

        # Load excel metadata
        wb = load_workbook(filename=xl_path)
        sheet = wb.active

        self.samples = []

        # rows start at min_row=2 to skip header
        for row in sheet.iter_rows(min_row=2, values_only=True):
            path, letter, dist = row
            self.samples.append((path, letter, float(dist)))

        # A → 0, B → 1, ..., Z → 25
        self.letter_to_idx = {chr(i + 65): i for i in range(26)}

    # ------------------------------------------------------------------
    # __len__
    # ------------------------------------------------------------------
    """
    Returns:
        int: Total number of samples in the dataset.
    """
    def __len__(self):
        return len(self.samples)

    # ------------------------------------------------------------------
    # __getitem__
    # ------------------------------------------------------------------
    """
    Retrieves and processes a single dataset entry.

    Args:
        idx (int): Index of the sample.

    Returns:
        (Tensor, int, Tensor):
            img_tensor : Float32 tensor, shape (3,128,128), normalized to [0,1]
            class_idx  : Integer label (0–25)
            dist       : 1D tensor containing the distance value

    Steps:
        1. Resolve image path from metadata.
        2. Load the image via PIL and convert to RGB.
        3. Resize to 128×128.
        4. Convert raw bytes → Tensor of shape (3, H, W).
        5. Normalize to [0,1].
        6. Convert letter → class index.
        7. Convert distance → float tensor.
    """
    def __getitem__(self, idx):
        rel_path, letter, dist = self.samples[idx]
        img_path = os.path.join(self.root, rel_path)

        # Load + preprocess the image
        img = Image.open(img_path).convert("RGB")
        img = img.resize((128, 128))

        # Convert PIL image → torch tensor
        img_tensor = torch.tensor(
            (torch.ByteTensor(torch.ByteStorage.from_buffer(img.tobytes()))
             .float()
             .view(128, 128, 3)
             .permute(2, 0, 1)) / 255.0
        )

        # Class label
        class_idx = self.letter_to_idx[letter]

        # Distance regression target
        dist = torch.tensor([dist], dtype=torch.float32)

        return img_tensor, class_idx, dist
